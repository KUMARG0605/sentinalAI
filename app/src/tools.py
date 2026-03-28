import json
import os
import shutil
import subprocess
import time
import webbrowser
import csv
from datetime import datetime
from pathlib import Path
from typing import Any, Callable
from urllib.parse import quote_plus

from langchain_core.documents import Document
from langchain_core.tools import tool

_RETRIEVER = None

def _find_vscode_exe() -> str:
    """Find VSCode exe — checks registry, known paths, then PATH."""
    import os as _os
    candidates = [
        _os.path.expandvars(r"%LOCALAPPDATA%\Programs\Microsoft VS Code\Code.exe"),
        r"C:\Program Files\Microsoft VS Code\Code.exe",
        r"C:\Program Files (x86)\Microsoft VS Code\Code.exe",
    ]
    for c in candidates:
        if _os.path.exists(c):
            return c
    # Fall back to PATH-based shutil.which
    import shutil as _sh
    return _sh.which("code") or _sh.which("code.exe") or "code.exe"


APP_OPEN_MAP: dict[str, str] = {
    # Browsers
    "chrome":       "chrome.exe",
    "firefox":      "firefox.exe",
    "edge":         "msedge.exe",
    # Editors / IDEs
    "notepad":      "notepad.exe",
    "notepad++":    "notepad++.exe",
    "vscode":       _find_vscode_exe(),
    "vs code":      _find_vscode_exe(),
    "visual studio code": _find_vscode_exe(),
    "word":         "winword.exe",
    "excel":        "excel.exe",
    "powerpoint":   "powerpnt.exe",
    # System
    "calculator":   "calc.exe",
    "calc":         "calc.exe",
    "cmd":          "cmd.exe",
    "command prompt": "cmd.exe",
    "terminal":     "wt.exe",
    "powershell":   "powershell.exe",
    "task manager": "taskmgr.exe",
    "explorer":     "explorer.exe",
    "file explorer": "explorer.exe",
    # Media
    "vlc":          "vlc.exe",
    "spotify":      "spotify.exe",
    "windows media player": "wmplayer.exe",
    # Messaging / Social
    "whatsapp":     "WhatsApp.exe",
    "telegram":     "telegram.exe",
    "discord":      "discord.exe",
    "slack":        "slack.exe",
    "teams":        "teams.exe",
    "microsoft teams": "teams.exe",
    "zoom":         "zoom.exe",
    # Utilities
    "paint":        "mspaint.exe",
    "snipping tool": "snippingtool.exe",
    "settings":     "ms-settings:",
    "control panel": "control.exe",
}

APP_PROCESS_MAP: dict[str, str] = {
    "chrome":       "chrome.exe",
    "firefox":      "firefox.exe",
    "edge":         "msedge.exe",
    "notepad":      "notepad.exe",
    "notepad++":    "notepad++.exe",
    "vscode":       "Code.exe",
    "vs code":      "Code.exe",
    "visual studio code": "Code.exe",
    "word":         "WINWORD.EXE",
    "excel":        "EXCEL.EXE",
    "powerpoint":   "POWERPNT.EXE",
    "calculator":   "CalculatorApp.exe",
    "calc":         "CalculatorApp.exe",
    "cmd":          "cmd.exe",
    "command prompt": "cmd.exe",
    "terminal":     "WindowsTerminal.exe",
    "powershell":   "powershell.exe",
    "task manager": "Taskmgr.exe",
    "explorer":     "explorer.exe",
    "file explorer": "explorer.exe",
    "vlc":          "vlc.exe",
    "spotify":      "Spotify.exe",
    "whatsapp":     "WhatsApp.exe",
    "telegram":     "Telegram.exe",
    "discord":      "Discord.exe",
    "slack":        "slack.exe",
    "teams":        "Teams.exe",
    "zoom":         "Zoom.exe",
    "paint":        "mspaint.exe",
}

MEDIA_EXTENSIONS = {".mp3", ".wav", ".flac", ".aac", ".m4a", ".mp4", ".mkv", ".avi", ".mov"}


def set_retriever(retriever) -> None:
    global _RETRIEVER
    _RETRIEVER = retriever


def _safe_text(value: Any) -> str:
    return str(value or "").strip()


def _resolve_app_command(app_name: str) -> str:
    key = _safe_text(app_name).lower()
    return APP_OPEN_MAP.get(key, app_name)


def _resolve_process_name(app_name: str) -> str:
    key = _safe_text(app_name).lower()
    return APP_PROCESS_MAP.get(key, app_name)


def _get_pyautogui():
    try:
        import pyautogui
        pyautogui.FAILSAFE = True
        return pyautogui
    except Exception as exc:
        raise RuntimeError(f"pyautogui is required for GUI tools: {exc}") from exc


def _get_pywinauto():
    try:
        from pywinauto import Application, Desktop, keyboard
        return Application, Desktop, keyboard
    except Exception as exc:
        raise RuntimeError(f"pywinauto is required for window automation tools: {exc}") from exc


def _get_windows_by_title(title_regex: str, timeout: float = 5.0) -> list:
    """
    Find windows matching a title regex with a hard timeout.

    Uses Win32 EnumWindows directly instead of pywinauto Desktop().windows()
    to avoid hanging on apps with complex UIA trees (VSCode, Chrome, Electron apps).
    Falls back to pywinauto for click/type operations on matched windows.
    """
    import ctypes, ctypes.wintypes as wt, re as _re

    compiled = _re.compile(title_regex, _re.IGNORECASE)
    matches = []

    @ctypes.WINFUNCTYPE(wt.BOOL, wt.HWND, wt.LPARAM)
    def enum_proc(hwnd, lparam):
        if ctypes.windll.user32.IsWindowVisible(hwnd):
            length = ctypes.windll.user32.GetWindowTextLengthW(hwnd)
            if length > 0:
                buf = ctypes.create_unicode_buffer(length + 1)
                ctypes.windll.user32.GetWindowTextW(hwnd, buf, length + 1)
                if compiled.search(buf.value):
                    matches.append(hwnd)
        return True

    ctypes.windll.user32.EnumWindows(enum_proc, 0)
    return matches


class _FastWindow:
    """
    Lightweight window wrapper that uses Win32 directly for fast operations
    and only falls back to pywinauto UIA for complex interactions.
    """
    def __init__(self, hwnd: int):
        import ctypes
        self.handle = hwnd
        length = ctypes.windll.user32.GetWindowTextLengthW(hwnd)
        buf = ctypes.create_unicode_buffer(length + 1)
        ctypes.windll.user32.GetWindowTextW(hwnd, buf, length + 1)
        self._title = buf.value

    def window_text(self) -> str:
        return self._title

    def set_focus(self) -> None:
        """Bring window to foreground (only call when absolutely needed)."""
        import ctypes
        ctypes.windll.user32.SetForegroundWindow(self.handle)
        import time; time.sleep(0.2)

    def rectangle(self):
        """Return RECT for the window."""
        import ctypes, ctypes.wintypes as wt
        rect = wt.RECT()
        ctypes.windll.user32.GetWindowRect(self.handle, ctypes.byref(rect))
        return rect

    def _get_pywinauto_win(self, timeout: float = 3.0):
        """Get pywinauto wrapper for this window (with timeout)."""
        import threading, queue as _q
        result_q = _q.Queue()
        def _fetch():
            try:
                from pywinauto import Desktop
                wins = Desktop(backend="uia").windows(handle=self.handle)
                result_q.put(wins[0] if wins else None)
            except Exception as e:
                result_q.put(None)
        t = threading.Thread(target=_fetch, daemon=True)
        t.start()
        try:
            return result_q.get(timeout=timeout)
        except _q.Empty:
            return None

    def child_window(self, title="", control_type=""):
        """Find a child control by title/type."""
        from app.src.background_actions import get_child_by_text
        hwnd = get_child_by_text(self.handle, title, control_type)
        if hwnd:
            return _FastWindow(hwnd)
        return None

    def click_input(self) -> None:
        """Click the center of this window."""
        import ctypes, ctypes.wintypes as wt
        rect = self.rectangle()
        cx = (rect.left + rect.right) // 2
        cy = (rect.top + rect.bottom) // 2
        from app.src.background_actions import click_background
        click_background(self.handle, cx, cy)


def _search_media_file(query: str, return_all: bool = False):
    """Search for media files. Returns first match (Path) or list of all matches if return_all=True."""
    candidate = Path(query).expanduser()
    if candidate.exists() and candidate.is_file():
        return [candidate] if return_all else candidate

    # Standard Windows media library paths
    home = Path.home()
    roots = [
        home / "Music",
        home / "Videos",
        home / "Downloads",
        home / "Desktop",
        home / "OneDrive" / "Music",
        home / "OneDrive" / "Videos",
        Path("C:/Users/Public/Music"),
        Path("C:/Users/Public/Videos"),
    ]
    query_lower = query.lower().strip()
    found = []

    for root in roots:
        if not root.exists():
            continue
        try:
            for path in root.rglob("*"):
                if not path.is_file():
                    continue
                if path.suffix.lower() not in MEDIA_EXTENSIONS:
                    continue
                # Match by query keyword or return any media file if query is generic
                if not query_lower or query_lower in path.name.lower():
                    if return_all:
                        found.append(path)
                        if len(found) >= 50:  # cap to avoid huge lists
                            return found
                    else:
                        return path
        except PermissionError:
            continue

    if return_all:
        return found
    return None


def _resolve_special_folder(folder_name: str) -> Path:
    """Return the true path for a Windows special folder (Desktop, Documents, etc).
    Handles OneDrive redirection."""
    import ctypes.wintypes
    CSIDL_MAP = {
        "desktop": 0x0000,
        "documents": 0x0005,
        "pictures": 0x0027,
        "music": 0x000D,
        "videos": 0x000E,
    }
    csidl = CSIDL_MAP.get(folder_name.lower())
    if csidl is None:
        return Path.home() / folder_name.title()

    buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
    ctypes.windll.shell32.SHGetFolderPathW(None, csidl, None, 0, buf)
    if buf.value:
        return Path(buf.value)
    return Path.home() / folder_name.title()

def _expand_path(path_str: str) -> Path:
    """Expand ~ and resolve common hardcoded C:/Users/name/Desktop paths to the real ones."""
    p_str = path_str.replace("\\", "/")
    
    # Auto-fix hardcoded desktop/document paths that might be redirected by OneDrive
    home_str = str(Path.home()).replace("\\", "/")
    if p_str.lower().startswith(home_str.lower() + "/desktop"):
        real_desk = str(_resolve_special_folder("desktop")).replace("\\", "/")
        p_str = real_desk + p_str[int(len(home_str + "/desktop")):]
    elif p_str.lower().startswith(home_str.lower() + "/documents"):
        real_docs = str(_resolve_special_folder("documents")).replace("\\", "/")
        p_str = real_docs + p_str[int(len(home_str + "/documents")):]

    return Path(p_str).expanduser()

# ── Knowledge base ────────────────────────────────────────────────────────────

def _tool_search_knowledge_base(query: str) -> str:
    if _RETRIEVER is None:
        return (
            "RAG_UNAVAILABLE: The knowledge base is not loaded. "
            "No FAISS index exists yet — it must be built first using build_index.py. "
            "Do NOT answer this question from general knowledge. "
            "Tell the user: 'The knowledge base is not set up yet. "
            "Please run build_index.py to index your documents first.'"
        )
    docs: list[Document] = _RETRIEVER.invoke(query)
    if not docs:
        return "RAG_NO_RESULTS: The knowledge base is loaded but no relevant documents were found for this query."
    blocks = []
    for i, doc in enumerate(docs[:3], 1):
        source = doc.metadata.get("source", "unknown")
        page = doc.metadata.get("page", "unknown")
        snippet = doc.page_content.strip().replace("\n", " ")
        if len(snippet) > 1200:
            snippet = snippet[:1200] + "..."
        blocks.append(f"[{i}] source={source} page={page}\n{snippet}")
    return "RAG_RESULTS:\n" + "\n\n".join(blocks)


# ── Application control ───────────────────────────────────────────────────────

def _wait_for_window(title_re: str, timeout: float = 5.0, Desktop=None):
    """Poll until exactly one window matching title_re exists. Returns the window or raises."""
    import re
    deadline = time.time() + timeout
    while time.time() < deadline:
        handles = _get_windows_by_title(title_re)
        if handles:
            return _FastWindow(max(handles))
        time.sleep(0.3)
    raise TimeoutError(f"No window matching '{title_re}' appeared within {timeout}s")


def _kill_all(process_name: str) -> None:
    """Kill all instances of a process silently."""
    subprocess.run(["taskkill", "/IM", process_name, "/F"],
                   capture_output=True, text=True)
    time.sleep(0.5)


def _tool_open_application(app_name: str, command: str = "", wait: float = 1.5) -> str:
    """Launch an application. Uses os.startfile to handle paths with spaces correctly."""
    from app.src.app_registry import registry as _reg

    if command.strip():
        exe = command.strip()
    else:
        exe = _reg.get(app_name) or _resolve_app_command(app_name)

    if not exe:
        return f"Could not find executable for '{app_name}'"

    # Resolve exe to full path if it's just a name like "code.exe"
    if exe and not os.path.isabs(exe):
        # Try to find full path via shutil.which (checks PATH)
        full = shutil.which(exe) or shutil.which(exe.replace(".exe", ""))
        if full:
            exe = full
        else:
            # Try app registry again with normalized name
            from app.src.app_registry import registry as _reg2
            found = _reg2.get(exe.replace(".exe", ""))
            if found and os.path.exists(found):
                exe = found

    # os.startfile works for full paths; subprocess shell=True for anything else
    if exe and os.path.exists(exe):
        try:
            os.startfile(exe)
            time.sleep(wait)
            return f"Opened: {app_name} ({exe})"
        except Exception:
            pass

    # Fallback: shell=True handles PATH-based commands and quoted paths with spaces
    if exe:
        try:
            cmd = f'"{exe}"' if os.path.exists(exe) else exe
            subprocess.Popen(cmd, shell=True,
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            time.sleep(wait)
            return f"Opened: {app_name} ({exe})"
        except Exception as exc:
            return (
                f"Could not open '{app_name}'. Path tried: '{exe}'. "
                f"Call find_app_path('{app_name}') to locate it, or "
                f"register_app_path('{app_name}', 'C:/full/path/app.exe') to register it."
            )
    return f"Could not find executable for '{app_name}'. Try register_app_path('{app_name}', 'C:/full/path.exe')"

def _tool_open_file_with_app(file_path: str, app_name: str = "", wait: float = 1.5) -> str:
    path = Path(file_path).expanduser()
    if not path.exists():
        return f"File not found: {path}"
    if app_name.strip():
        from app.src.app_registry import registry as _reg
        app_cmd = _reg.get(app_name) or _resolve_app_command(app_name)
        try:
            subprocess.Popen(
                [app_cmd, str(path)],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                creationflags=subprocess.DETACHED_PROCESS if hasattr(subprocess, 'DETACHED_PROCESS') else 0,
            )
            time.sleep(wait)
            return f"Opened '{path}' with '{app_cmd}'"
        except Exception as exc:
            return f"Failed to open file with '{app_cmd}': {exc}"
    try:
        os.startfile(str(path))
        time.sleep(wait)
        return f"Opened file: {path}"
    except Exception as exc:
        return f"Failed to open file '{path}': {exc}"


def _tool_close_application(app_name: str, window_title_regex: str = "") -> str:
    process_name = _resolve_process_name(app_name)
    if window_title_regex.strip():
        try:
            import ctypes
            handles = _get_windows_by_title(window_title_regex)
            if handles:
                hwnd = max(handles)
                # WM_CLOSE = 0x0010
                ctypes.windll.user32.PostMessageW(hwnd, 0x0010, 0, 0)
                time.sleep(0.4)
                return f"Closed window matching: {window_title_regex}"
        except Exception:
            pass
    try:
        completed = subprocess.run(
            ["taskkill", "/IM", process_name, "/F"],
            capture_output=True, text=True,
        )
        time.sleep(0.4)
        if completed.returncode == 0:
            return f"Closed application/process: {process_name}"
        return f"Could not close '{process_name}'. Output: {completed.stderr or completed.stdout}"
    except Exception as exc:
        return f"Failed to close '{process_name}': {exc}"


def _tool_kill_app_instances(app_name: str) -> str:
    """Force-kill ALL instances of an app and wait for them to fully terminate."""
    process_name = _resolve_process_name(app_name)
    result = subprocess.run(["taskkill", "/IM", process_name, "/F"],
                            capture_output=True, text=True)
    time.sleep(0.8)
    # Verify gone
    check = subprocess.run(["tasklist", "/FI", f"IMAGENAME eq {process_name}"],
                           capture_output=True, text=True)
    if process_name.lower() in check.stdout.lower():
        return f"WARNING: Some '{process_name}' processes may still be running."
    return f"All instances of '{process_name}' terminated."


def _tool_sleep(seconds: float) -> str:
    """Pause execution for the given number of seconds."""
    secs = max(0.1, min(float(seconds), 30.0))
    time.sleep(secs)
    return f"Slept {secs:.1f}s."


def _tool_focus_window(window_title_regex: str) -> str:
    """
    Confirm a window exists. Uses fast Win32 EnumWindows — never hangs.
    Does NOT steal focus from the user's current app.
    """
    from app.src.background_actions import get_window_by_regex, get_window_title
    hwnd = get_window_by_regex(window_title_regex)
    if hwnd:
        title = get_window_title(hwnd)
        return f"Window confirmed: '{title}' (background, focus unchanged)"
    # Also try _get_windows_by_title as backup
    handles = _get_windows_by_title(window_title_regex)
    if handles:
        win = _FastWindow(max(handles))
        return f"Window confirmed: '{win.window_text()}' (background)"
    return f"No window found matching: '{window_title_regex}' — app may still be loading"

def _tool_click_window_control(window_title_regex: str, control_title: str = "", control_type: str = "") -> str:
    try:
        _, Desktop, _ = _get_pywinauto()
        # Fast Win32 lookup — never hangs on VSCode/Chrome/Electron apps
        _handles = _get_windows_by_title(window_title_regex)
        if not _handles:
            return f"No window found matching: '{window_title_regex}'"
        win = _FastWindow(max(_handles))
        win.set_focus()
        if control_title.strip():
            control = win.child_window(title=control_title, control_type=control_type or None)
            control.wrapper_object().click_input()
            return f"Clicked control '{control_title}' in window '{win.window_text()}'."
        win.click_input()
        return f"Clicked focused window '{win.window_text()}'."
    except Exception as exc:
        return f"Failed to click control/window '{window_title_regex}': {exc}"


def _tool_type_in_window(window_title_regex: str, text: str, with_enter: bool = False) -> str:
    try:
        _, Desktop, keyboard = _get_pywinauto()
        # Fast Win32 lookup — never hangs on VSCode/Chrome/Electron apps
        _handles = _get_windows_by_title(window_title_regex)
        if not _handles:
            return f"No window found matching: '{window_title_regex}'"
        win = _FastWindow(max(_handles))
        win.set_focus()
        time.sleep(0.3)

        # ── Normalise newlines ──────────────────────────────────────────────
        # The LLM may send literal \n (two chars) OR a real newline.
        # pywinauto send_keys needs {ENTER} for newlines.
        # Also escape pywinauto special chars: { } + ^ % ~ ( ) [ ]
        def _to_sendkeys(raw: str) -> str:
            # First replace literal backslash-n sequences that LLM sometimes sends
            raw = raw.replace("\\n", "\n")
            result = []
            for ch in raw:
                if ch == "\n":
                    result.append("{ENTER}")
                elif ch in r"{}+^%~()[]":
                    result.append("{" + ch + "}")
                else:
                    result.append(ch)
            return "".join(result)

        payload = _to_sendkeys(text)
        if with_enter:
            payload += "{ENTER}"

        keyboard.send_keys(payload, with_spaces=True, pause=0.02)
        return f"Typed text in window '{win.window_text()}'."
    except Exception as exc:
        return f"Failed typing in window '{window_title_regex}': {exc}"


# ── Keyboard / Mouse ──────────────────────────────────────────────────────────

def _tool_keyboard_type(text: str) -> str:
    try:
        pyautogui = _get_pyautogui()
        pyautogui.write(text, interval=0.01)
        return "Typed text successfully."
    except Exception as exc:
        return f"Keyboard typing failed: {exc}"


def _tool_keyboard_press(keys: str) -> str:
    try:
        pyautogui = _get_pyautogui()
        parts = [p.strip().lower() for p in keys.split("+") if p.strip()]
        if len(parts) > 1:
            pyautogui.hotkey(*parts)
            return f"Executed hotkey: {'+'.join(parts)}"
        if not parts:
            return "No key provided."
        pyautogui.press(parts[0])
        return f"Pressed key: {parts[0]}"
    except Exception as exc:
        return f"Keyboard key press failed: {exc}"


def _tool_mouse_move(x: int, y: int, duration: float = 0.2) -> str:
    try:
        pyautogui = _get_pyautogui()
        pyautogui.moveTo(x, y, duration=max(duration, 0.0))
        return f"Moved mouse to ({x}, {y})."
    except Exception as exc:
        return f"Mouse move failed: {exc}"


def _tool_mouse_click(button: str = "left", clicks: int = 1, x: int = -1, y: int = -1) -> str:
    try:
        pyautogui = _get_pyautogui()
        target_x = None if x < 0 else x
        target_y = None if y < 0 else y
        pyautogui.click(x=target_x, y=target_y, button=button, clicks=max(clicks, 1))
        pos = f"at ({target_x}, {target_y})" if target_x is not None else "at current position"
        return f"Clicked mouse {pos}, button={button}, clicks={clicks}."
    except Exception as exc:
        return f"Mouse click failed: {exc}"


# ── Screenshot ────────────────────────────────────────────────────────────────

def _tool_take_screenshot(file_path: str = "") -> str:
    try:
        pyautogui = _get_pyautogui()
        if not file_path.strip():
            shots_dir = Path.cwd() / "data" / "screenshots"
            shots_dir.mkdir(parents=True, exist_ok=True)
            file_path = str(shots_dir / f"shot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        output = Path(file_path).expanduser()
        output.parent.mkdir(parents=True, exist_ok=True)
        image = pyautogui.screenshot()
        image.save(str(output))
        return f"Screenshot saved: {output}"
    except Exception as exc:
        return f"Screenshot failed: {exc}"


# ── Vision Desktop Actions ───────────────────────────────────────────────────

def _tool_vision_act_on_screen(element_description: str, action: str = "click", type_text: str = "") -> str:
    try:
        from PIL import ImageDraw, ImageFont
        import pyautogui
        import base64
        import json
        import os
        from app.src.llm_rotation import get_llm
        from langchain_core.messages import HumanMessage
    except ImportError as e:
        return f"Import failed (run pip install pyautogui pillow langchain-sambanova langchain-core): {e}"

    try:
        # Take screenshot
        img = pyautogui.screenshot()
        width, height = img.size
        
        # Grid settings (20 cols x 10 rows)
        cols, rows = 20, 10
        cell_w, cell_h = width / cols, height / rows
        
        # Draw grid
        draw = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype("arial.ttf", size=max(12, int(cell_h / 3)))
        except IOError:
            font = ImageFont.load_default()
            
        # Draw lines
        for r in range(rows + 1):
            y = int(r * cell_h)
            draw.line([(0, y), (width, y)], fill=(255, 0, 0, 128), width=2)
        for c in range(cols + 1):
            x = int(c * cell_w)
            draw.line([(x, 0), (x, height)], fill=(255, 0, 0, 128), width=2)
            
        # Draw numbers in center of each cell
        cell_centers = {}
        idx = 1
        for r in range(rows):
            for c in range(cols):
                cx, cy = int((c + 0.5) * cell_w), int((r + 0.5) * cell_h)
                cell_centers[idx] = (cx, cy)
                
                # Outline effect for text visibility
                txt = str(idx)
                for ox, oy in [(-1,-1), (-1,1), (1,-1), (1,1)]:
                    draw.text((cx-10+ox, cy-10+oy), txt, fill="black", font=font)
                draw.text((cx-10, cy-10), txt, fill="white", font=font)
                idx += 1
                
        # Convert to base64
        import io
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        b64_img = base64.b64encode(buf.getvalue()).decode()
        
        # Ask Vision Model
        model_name = os.environ.get("SAMBANOVA_VISION_MODEL", "Llama-3.2-90B-Vision-Instruct")
        llm = get_llm(model=model_name, temperature=0.1)
        
        msg = HumanMessage(content=[
            {"type": "text", "text": (
                f"You are a desktop automation agent controlling a computer. "
                f"I have overlaid a {cols}x{rows} numbered grid on the screen.\n\n"
                f"Find the element described as: '{element_description}'.\n"
                f"Which grid cell number (from 1 to {cols*rows}) is directly over the center of this element?\n\n"
                "Return YOUR RESPONSE AS A VALID JSON OBJECT ONLY. Example format:\n"
                "{\"cell_number\": 45}\n"
                "Do NOT include any extra text."
            )},
            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_img}"}},
        ])
        
        resp = llm.invoke([msg])
        reply = str(resp.content).strip()
        
        # Parse JSON
        if "```json" in reply:
            reply = reply.split("```json")[-1].split("```")[0].strip()
        elif "```" in reply:
            reply = reply.split("```")[-1].split("```")[0].strip()
            
        try:
            data = json.loads(reply)
            cell = int(data.get("cell_number", -1))
        except Exception:
            import re
            nums = re.findall(r'\\b\\d+\\b', reply)
            if not nums:
                return f"Vision Model failed to return a valid cell number. Raw output: {resp.content}"
            cell = int(nums[-1])
            
        if cell not in cell_centers:
            return f"Invalid cell number returned ({cell}). Must be between 1 and {cols*rows}."
            
        # Perform action
        cx, cy = cell_centers[cell]
        
        act = action.lower()
        if act in ["click", "left_click"]:
            pyautogui.click(cx, cy)
            res = f"Clicked cell {cell} at coordinates ({cx}, {cy})."
        elif act == "double_click":
            pyautogui.doubleClick(cx, cy)
            res = f"Double-clicked cell {cell} at ({cx}, {cy})."
        elif act == "right_click":
            pyautogui.rightClick(cx, cy)
            res = f"Right-clicked cell {cell} at ({cx}, {cy})."
        elif act == "click_and_type":
            pyautogui.click(cx, cy)
            import time; time.sleep(0.5)
            pyautogui.write(type_text, interval=0.01)
            res = f"Clicked cell {cell} at ({cx}, {cy}) and typed '{type_text[:20]}...'."
        else:
            return f"Cell {cell} identified at ({cx}, {cy}), but action '{action}' is invalid."
            
        return res
        
    except Exception as exc:
        return f"vision_act_on_screen failed: {exc}"


# ── Media ─────────────────────────────────────────────────────────────────────

def _tool_play_media(query_or_path: str) -> str:
    target = _search_media_file(query_or_path)
    if target is None:
        # If query is generic ("play music", "any song", etc.), grab the first file found
        generic_terms = {"music", "song", "audio", "video", "media", "any", "one", "file", "them"}
        if any(t in query_or_path.lower() for t in generic_terms):
            all_files = _search_media_file("", return_all=True)
            target = all_files[0] if all_files else None
    if target is None:
        return (
            f"No media files found for: '{query_or_path}'. "
            "Try find_media_files() to list what's available, then play a specific one."
        )
    try:
        os.startfile(str(target))
        return f"Playing: {target.name}  (path: {target})"
    except Exception as exc:
        return f"Failed to play media '{target}': {exc}"


def _tool_find_media_files(directory: str = "", extension_filter: str = "") -> str:
    exts = MEDIA_EXTENSIONS
    if extension_filter.strip():
        exts = {f".{e.strip().lstrip('.')}" for e in extension_filter.split(",")}

    if directory.strip():
        roots = [Path(directory).expanduser()]
    else:
        home = Path.home()
        roots = [
            home / "Music", home / "Videos", home / "Downloads",
            home / "Desktop",
            home / "OneDrive" / "Music", home / "OneDrive" / "Videos",
            Path("C:/Users/Public/Music"), Path("C:/Users/Public/Videos"),
        ]

    found = []
    for root in roots:
        if not root.exists():
            continue
        try:
            for path in root.rglob("*"):
                if path.is_file() and path.suffix.lower() in exts:
                    found.append(str(path))
                    if len(found) >= 50:
                        break
        except PermissionError:
            continue

    if not found:
        return (
            "No media files found in standard locations (Music, Videos, Downloads, Desktop, OneDrive). "
            "Provide a specific directory path to search elsewhere."
        )
    lines = [f"{i+1}. {p}" for i, p in enumerate(found)]
    return f"Found {len(found)} media file(s):\n" + "\n".join(lines)


# ── Web ───────────────────────────────────────────────────────────────────────

def _tool_web_search(query: str) -> str:
    url = f"https://www.google.com/search?q={quote_plus(query)}"
    webbrowser.open(url, new=2)
    return f"Opened web search for: {query}"


def _tool_open_website(url: str) -> str:
    clean = url.strip()
    if not clean:
        return "No URL provided."
    if not clean.startswith(("http://", "https://")):
        clean = "https://" + clean
    webbrowser.open(clean, new=2)
    return f"Opened website: {clean}"


# ── Advanced filesystem operations ───────────────────────────────────────────

def _tool_search_files(
    name_pattern: str = "",
    search_dir: str = "",
    extension: str = "",
    content_keyword: str = "",
    min_size_kb: float = -1,
    max_size_kb: float = -1,
    max_results: int = 50,
) -> str:
    """
    Recursively search for files/folders.
    Filters can be combined: name_pattern AND extension AND content_keyword AND size range.
    """
    import fnmatch

    root = Path(search_dir).expanduser() if search_dir.strip() else Path.home()
    if not root.exists():
        return f"Search root not found: {root}"

    ext_filter = extension.strip().lstrip(".").lower() if extension.strip() else ""
    pattern    = name_pattern.strip().lower()
    keyword    = content_keyword.strip().lower()
    min_bytes  = int(min_size_kb * 1024) if min_size_kb >= 0 else -1
    max_bytes  = int(max_size_kb * 1024) if max_size_kb >= 0 else -1

    matches: list[str] = []
    skipped_dirs = {"$recycle.bin", "system volume information", "windows", "appdata"}

    try:
        for entry in root.rglob("*"):
            if len(matches) >= max_results:
                break
            # Skip noisy system dirs
            if any(part.lower() in skipped_dirs for part in entry.parts):
                continue
            try:
                # Name pattern filter
                if pattern and not fnmatch.fnmatch(entry.name.lower(), f"*{pattern}*"):
                    continue
                # Extension filter (applies to files only)
                if ext_filter and entry.is_file():
                    if entry.suffix.lower().lstrip(".") != ext_filter:
                        continue
                # Size filter (files only)
                if entry.is_file() and (min_bytes >= 0 or max_bytes >= 0):
                    sz = entry.stat().st_size
                    if min_bytes >= 0 and sz < min_bytes:
                        continue
                    if max_bytes >= 0 and sz > max_bytes:
                        continue
                # Content keyword filter (text files only)
                if keyword and entry.is_file():
                    try:
                        text = entry.read_text(encoding="utf-8", errors="ignore")
                        if keyword not in text.lower():
                            continue
                    except Exception:
                        continue
                matches.append(str(entry))
            except PermissionError:
                continue
    except PermissionError:
        pass

    if not matches:
        return (
            f"No matches found in '{root}' for: "
            f"name='{name_pattern}' ext='{extension}' keyword='{content_keyword}'"
        )
    note = f" (showing first {max_results})" if len(matches) == max_results else ""
    lines = [f"{i+1}. {p}" for i, p in enumerate(matches)]
    return f"Found {len(matches)} result(s){note} in '{root}':\n" + "\n".join(lines)


def _tool_get_file_info(path: str) -> str:
    """Return detailed metadata for a file or folder."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"Path not found: {p}"

    try:
        stat = p.stat()
    except PermissionError:
        return f"Permission denied: {p}"

    def _fmt_bytes(n: int) -> str:
        for unit in ("B", "KB", "MB", "GB", "TB"):
            if n < 1024 or unit == "TB":
                return f"{n:.2f} {unit}" if unit != "B" else f"{n} B"
            n /= 1024

    def _ts(t: float) -> str:
        return datetime.fromtimestamp(t).strftime("%Y-%m-%d %H:%M:%S")

    lines = [
        f"Path       : {p}",
        f"Type       : {'Directory' if p.is_dir() else 'File'}",
        f"Name       : {p.name}",
        f"Extension  : {p.suffix or '(none)'}",
        f"Parent     : {p.parent}",
        f"Created    : {_ts(stat.st_ctime)}",
        f"Modified   : {_ts(stat.st_mtime)}",
        f"Accessed   : {_ts(stat.st_atime)}",
    ]

    if p.is_file():
        lines.append(f"Size       : {_fmt_bytes(stat.st_size)} ({stat.st_size:,} bytes)")
        # MIME type guess
        import mimetypes
        mime, _ = mimetypes.guess_type(str(p))
        lines.append(f"MIME type  : {mime or 'unknown'}")
        # Is readable as text?
        try:
            p.read_text(encoding="utf-8", errors="strict")
            lines.append("Encoding   : UTF-8 text")
        except UnicodeDecodeError:
            lines.append("Encoding   : Binary")
        except PermissionError:
            lines.append("Encoding   : (permission denied)")
    else:
        # Directory stats
        try:
            children = list(p.iterdir())
            files   = sum(1 for c in children if c.is_file())
            subdirs = sum(1 for c in children if c.is_dir())
            lines.append(f"Contents   : {files} file(s), {subdirs} subdirectory(-ies) (immediate children)")
            # Total size (non-recursive for speed)
            total = sum(c.stat().st_size for c in children if c.is_file())
            lines.append(f"Total size : {_fmt_bytes(total)} (immediate files only)")
        except PermissionError:
            lines.append("Contents   : (permission denied)")

    # Read-only flag
    import stat as stat_mod
    readable  = bool(stat.st_mode & stat_mod.S_IRUSR)
    writable  = bool(stat.st_mode & stat_mod.S_IWUSR)
    lines.append(f"Readable   : {readable}")
    lines.append(f"Writable   : {writable}")

    return "\n".join(lines)


def _tool_list_folder_tree(
    directory_path: str,
    max_depth: int = 3,
    max_items: int = 200,
) -> str:
    """Return an indented tree of a folder's contents up to max_depth levels."""
    root = Path(directory_path).expanduser()
    if not root.exists() or not root.is_dir():
        return f"Directory not found: {root}"

    lines: list[str] = [str(root)]
    count = [0]

    def _walk(path: Path, depth: int, prefix: str) -> None:
        if depth > max_depth or count[0] >= max_items:
            return
        try:
            entries = sorted(path.iterdir(), key=lambda e: (e.is_file(), e.name.lower()))
        except PermissionError:
            lines.append(prefix + "  [permission denied]")
            return
        for i, entry in enumerate(entries):
            if count[0] >= max_items:
                lines.append(prefix + "  ... [truncated]")
                return
            connector = "└── " if i == len(entries) - 1 else "├── "
            icon = "📁 " if entry.is_dir() else "📄 "
            size_hint = ""
            if entry.is_file():
                try:
                    sz = entry.stat().st_size
                    if sz >= 1_048_576:
                        size_hint = f" ({sz/1_048_576:.1f} MB)"
                    elif sz >= 1024:
                        size_hint = f" ({sz/1024:.1f} KB)"
                    else:
                        size_hint = f" ({sz} B)"
                except OSError:
                    pass
            lines.append(f"{prefix}{connector}{icon}{entry.name}{size_hint}")
            count[0] += 1
            if entry.is_dir():
                extension_prefix = prefix + ("    " if i == len(entries) - 1 else "│   ")
                _walk(entry, depth + 1, extension_prefix)

    _walk(root, 1, "")
    if count[0] >= max_items:
        lines.append(f"\n[Showing first {max_items} items. Use a sub-path to see more.]")
    return "\n".join(lines)


# ── File operations ───────────────────────────────────────────────────────────

def _write_file_via_cmd(file_path: str, content: str, mode: str = "w") -> str:
    """
    Write or append text to a file using PowerShell Set-Content / Add-Content.
    This bypasses Windows file-locking issues that occur when apps like Notepad
    hold the file open (Python's open() raises PermissionError in that case).

    mode: 'w' = overwrite, 'a' = append
    """
    path = _expand_path(file_path)
    file_path = str(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # Escape single-quotes in content for PowerShell
    escaped = content.replace("'", "''")

    if mode == "a":
        ps_cmd = f"Add-Content -Path '{path}' -Value '{escaped}' -Encoding UTF8"
    else:
        ps_cmd = f"Set-Content -Path '{path}' -Value '{escaped}' -Encoding UTF8"

    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_cmd],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            err = result.stderr.strip()
            # Fallback to direct Python write if PowerShell fails
            flag = "a" if mode == "a" else "w"
            with path.open(flag, encoding="utf-8") as f:
                f.write(content)
            return f"{'Appended' if mode == 'a' else 'Created'} (py-fallback): {path}"
        action = "Appended to" if mode == "a" else "Created"
        return f"{action} file: {path}"
    except Exception as exc:
        # Last-resort direct write
        flag = "a" if mode == "a" else "w"
        with path.open(flag, encoding="utf-8") as f:
            f.write(content)
        return f"{'Appended' if mode == 'a' else 'Created'} (py-fallback): {path}"


def _tool_create_file(file_path: str, content: str = "") -> str:
    return _write_file_via_cmd(file_path, content, mode="w")


def _tool_append_file(file_path: str, content: str) -> str:
    path = _expand_path(file_path)
    if not path.exists():
        return f"File not found: {path}. Use create_file to create it first."
    return _write_file_via_cmd(file_path, content, mode="a")


def _tool_read_file(file_path: str) -> str:
    path = _expand_path(file_path)
    if not path.exists() or not path.is_file():
        return f"File not found: {path}"
    try:
        content = path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return f"File is not UTF-8 text: {path}"
    if len(content) > 4000:
        content = content[:4000] + "\n... [truncated]"
    return content


def _tool_delete_file(file_path: str) -> str:
    path = _expand_path(file_path)
    if not path.exists():
        return f"File not found: {path}"
    try:
        path.unlink()
        return f"Deleted file: {path}"
    except Exception as exc:
        return f"Failed to delete '{path}': {exc}"


def _tool_copy_file(source_path: str, destination_path: str) -> str:
    src = _expand_path(source_path)
    dst = _expand_path(destination_path)
    if not src.exists():
        return f"Source file not found: {src}"
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(src), str(dst))
    return f"Copied '{src}' to '{dst}'"


def _tool_list_files(directory_path: str) -> str:
    directory = _expand_path(directory_path)
    if not directory.exists() or not directory.is_dir():
        return f"Directory not found: {directory}"
    items = sorted(directory.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
    if not items:
        return f"Directory is empty: {directory}"
    lines = []
    for item in items[:200]:
        kind = "FILE" if item.is_file() else "DIR"
        lines.append(f"{kind}: {item.name}")
    if len(items) > 200:
        lines.append(f"... and {len(items) - 200} more")
    return "\n".join(lines)


def _tool_write_csv(file_path: str, headers_json: str = "", rows_json: str = "", append: bool = False) -> str:
    path = _expand_path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    try:
        headers = json.loads(headers_json) if headers_json.strip() else []
        rows = json.loads(rows_json) if rows_json.strip() else []
    except json.JSONDecodeError as exc:
        return f"Invalid JSON input for CSV: {exc}"

    if headers and not isinstance(headers, list):
        return "headers_json must be a JSON array of column names."
    if rows and not isinstance(rows, list):
        return "rows_json must be a JSON array of row arrays or row objects."

    mode = "a" if append else "w"
    write_header = (not path.exists()) or (not append)

    try:
        with path.open(mode, newline="", encoding="utf-8") as f:
            if rows and isinstance(rows[0], dict):
                fieldnames = headers or list(rows[0].keys())
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                if write_header:
                    writer.writeheader()
                for row in rows:
                    writer.writerow(row)
            else:
                writer = csv.writer(f)
                if headers and write_header:
                    writer.writerow(headers)
                for row in rows:
                    writer.writerow(row if isinstance(row, list) else [row])
        return f"CSV {'appended' if append else 'written'}: {path}"
    except Exception as exc:
        return f"Failed writing CSV '{path}': {exc}"


def _tool_read_csv(file_path: str, max_rows: int = 50) -> str:
    path = _expand_path(file_path)
    if not path.exists():
        return f"CSV file not found: {path}"

    try:
        with path.open("r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = []
            for idx, row in enumerate(reader):
                if idx >= max_rows:
                    rows.append(["... [truncated]"])
                    break
                rows.append(row)
        return json.dumps(rows, ensure_ascii=False)
    except Exception as exc:
        return f"Failed reading CSV '{path}': {exc}"


def _tool_csv_to_excel(csv_path: str, excel_path: str, sheet_name: str = "Sheet1") -> str:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        return f"openpyxl is required for Excel tools: {exc}"

    src = _expand_path(csv_path)
    dst = _expand_path(excel_path)
    if not src.exists():
        return f"CSV file not found: {src}"
    dst.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        with src.open("r", newline="", encoding="utf-8") as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save(dst)
        return f"Converted CSV to Excel: {dst}"
    except Exception as exc:
        return f"CSV to Excel conversion failed: {exc}"


def _tool_excel_to_csv(excel_path: str, csv_path: str, sheet_name: str = "") -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return f"openpyxl is required for Excel tools: {exc}"

    src = _expand_path(excel_path)
    dst = _expand_path(csv_path)
    if not src.exists():
        return f"Excel file not found: {src}"
    dst.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(src, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        with dst.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])
        return f"Converted Excel to CSV: {dst}"
    except Exception as exc:
        return f"Excel to CSV conversion failed: {exc}"


def _tool_list_excel_sheets(excel_path: str) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return f"openpyxl is required for Excel tools: {exc}"

    src = _expand_path(excel_path)
    if not src.exists():
        return f"Excel file not found: {src}"

    try:
        wb = load_workbook(src, read_only=True, data_only=True)
        return json.dumps(wb.sheetnames, ensure_ascii=False)
    except Exception as exc:
        return f"Failed listing sheet names: {exc}"


def _tool_convert_image_format(input_path: str, output_format: str, output_path: str = "") -> str:
    try:
        from PIL import Image
    except Exception as exc:
        return f"Pillow is required for image conversion: {exc}"

    src = _expand_path(input_path)
    if not src.exists():
        return f"Image not found: {src}"

    fmt = output_format.strip().lower().lstrip(".")
    format_map = {
        "jpg": "JPEG",
        "jpeg": "JPEG",
        "png": "PNG",
        "webp": "WEBP",
        "bmp": "BMP",
        "tiff": "TIFF",
        "gif": "GIF",
    }
    if fmt not in format_map:
        return "Unsupported format. Use one of: jpg,jpeg,png,webp,bmp,tiff,gif"

    if output_path.strip():
        dst = _expand_path(output_path)
    else:
        dst = src.with_suffix(f".{fmt}")
    dst.parent.mkdir(parents=True, exist_ok=True)

    try:
        with Image.open(src) as img:
            if format_map[fmt] == "JPEG" and img.mode in ("RGBA", "LA", "P"):
                img = img.convert("RGB")
            img.save(dst, format=format_map[fmt])
        return f"Converted image: {src} -> {dst}"
    except Exception as exc:
        return f"Image conversion failed: {exc}"


# ── System utilities ──────────────────────────────────────────────────────────

def _tool_get_datetime() -> str:
    now = datetime.now()
    return now.strftime("Date: %A, %B %d, %Y | Time: %I:%M %p")


def _tool_run_shell_command(command: str) -> str:
    try:
        result = subprocess.run(
            command, shell=True, capture_output=True, text=True, timeout=30,
        )
        output = (result.stdout or "").strip()
        errors = (result.stderr or "").strip()
        if result.returncode != 0:
            return f"Command failed (exit {result.returncode}).\nSTDOUT: {output}\nSTDERR: {errors}"
        return output or "Command completed (exit 0, no output)."
    except subprocess.TimeoutExpired:
        return "Command timed out after 30 seconds."
    except Exception as exc:
        return f"Shell command failed: {exc}"


# ── Multi-step orchestrator ───────────────────────────────────────────────────

def _tool_run_multi_step_actions(steps_json: str) -> str:
    try:
        steps = json.loads(steps_json)
    except json.JSONDecodeError as exc:
        return f"Invalid JSON for steps: {exc}"

    if not isinstance(steps, list):
        return "Steps input must be a JSON array."

    actions: dict[str, Callable[..., str]] = {
        "search_knowledge_base": _tool_search_knowledge_base,
        "open_application": _tool_open_application,
        "open_in_app": _tool_open_in_app,
        "open_file_with_app": _tool_open_file_with_app,
        "close_application": _tool_close_application,
        "kill_app_instances": _tool_kill_app_instances,
        "sleep": _tool_sleep,
        "focus_window": _tool_focus_window,
        "click_window_control": _tool_click_window_control,
        "type_in_window": _tool_type_in_window,
        "keyboard_type": _tool_keyboard_type,
        "keyboard_press": _tool_keyboard_press,
        "mouse_move": _tool_mouse_move,
        "mouse_click": _tool_mouse_click,
        "take_screenshot": _tool_take_screenshot,
        "vision_act_on_screen": _tool_vision_act_on_screen,
        "play_media": _tool_play_media,
        "find_media_files": _tool_find_media_files,
        "web_search": _tool_web_search,
        "open_website": _tool_open_website,
        "create_file": _tool_create_file,
        "append_file": _tool_append_file,
        "read_file": _tool_read_file,
        "delete_file": _tool_delete_file,
        "copy_file": _tool_copy_file,
        "list_files": _tool_list_files,
        "write_csv": _tool_write_csv,
        "read_csv": _tool_read_csv,
        "csv_to_excel": _tool_csv_to_excel,
        "excel_to_csv": _tool_excel_to_csv,
        "list_excel_sheets": _tool_list_excel_sheets,
        "convert_image_format": _tool_convert_image_format,
        "search_files": _tool_search_files,
        "get_file_info": _tool_get_file_info,
        "list_folder_tree": _tool_list_folder_tree,
        "get_desktop_path": lambda: str(_resolve_special_folder("desktop")),
        "get_datetime": _tool_get_datetime,
        "run_shell_command": _tool_run_shell_command,
    }

    results = []
    for idx, step in enumerate(steps, 1):
        if not isinstance(step, dict):
            results.append(f"Step {idx}: invalid step format (expected object)")
            continue
        action = _safe_text(step.get("action"))
        args = step.get("args", {})
        if action not in actions:
            results.append(f"Step {idx}: unknown action '{action}'. Available: {sorted(actions)}")
            continue
        if not isinstance(args, dict):
            results.append(f"Step {idx}: args must be a JSON object")
            continue
        try:
            result = actions[action](**args)
        except TypeError as exc:
            result = f"argument error for '{action}': {exc}"
        except Exception as exc:
            result = f"action '{action}' failed: {exc}"
        results.append(f"Step {idx} ({action}): {result}")
    return "\n".join(results)


# ── LangChain @tool wrappers ──────────────────────────────────────────────────

@tool
def search_knowledge_base(query: str) -> str:
    """Search the local FAISS knowledge base for stored documents matching a query.
    Call this FIRST for any question about stored documents, company info, or knowledge topics.
    Returns up to 3 matching snippets with source and page metadata.
    Example: search_knowledge_base(query='how to reset password')"""
    return _tool_search_knowledge_base(query)


@tool
def open_application(app_name: str, command: str = "") -> str:
    """Open a desktop application by its common name or a full executable path.
    Known names: 'notepad', 'chrome', 'calculator', 'vlc', 'spotify'.
    For other apps, provide the full path in the 'command' argument.
    Example: open_application(app_name='notepad')
    Example: open_application(app_name='myapp', command='C:/Apps/myapp.exe')"""
    return _tool_open_application(app_name, command)


@tool
def open_file_with_app(file_path: str, app_name: str = "") -> str:
    """Open a specific file with an application.
    Provide the full file path and optionally the app to open it with.
    If app_name is omitted, uses the system default for that file type.
    Example: open_file_with_app(file_path='C:/Users/user/doc.txt', app_name='notepad')
    Example: open_file_with_app(file_path='C:/Users/user/report.pdf')"""
    return _tool_open_file_with_app(file_path, app_name)


@tool
def close_application(app_name: str, window_title_regex: str = "") -> str:
    """Close a running application by name (e.g. 'notepad', 'chrome', 'vlc').
    Optionally provide window_title_regex to target a specific window.
    Example: close_application(app_name='notepad')
    Example: close_application(app_name='chrome', window_title_regex='.*Gmail.*')"""
    return _tool_close_application(app_name, window_title_regex)


@tool
def kill_app_instances(app_name: str) -> str:
    """Force-kill ALL running instances of an app and wait for full termination.
    Use this BEFORE open_application when you need a clean single instance (e.g. before opening Notepad
    to ensure there are no leftover windows that would cause 'multiple elements' errors).
    Example: kill_app_instances(app_name='notepad')"""
    return _tool_kill_app_instances(app_name)


@tool
def sleep(seconds: float) -> str:
    """Pause execution for a number of seconds (max 30). Use after open_application to let the
    window fully load before interacting with it, or between steps that need timing.
    Example: sleep(seconds=1.5)"""
    return _tool_sleep(seconds)


@tool
def focus_window(window_title_regex: str) -> str:
    """Bring a window to the foreground by matching its title using a regex pattern.
    Example: focus_window(window_title_regex='Notepad')
    Example: focus_window(window_title_regex='.*prompt.txt.*')"""
    return _tool_focus_window(window_title_regex)


@tool
def click_window_control(window_title_regex: str, control_title: str = "", control_type: str = "") -> str:
    """Click a UI control inside a window identified by title regex.
    Provide control_title (button/label text) and optionally control_type ('Button', 'Edit', etc.).
    Example: click_window_control(window_title_regex='Save As', control_title='Save', control_type='Button')"""
    return _tool_click_window_control(window_title_regex, control_title, control_type)


@tool
def type_in_window(window_title_regex: str, text: str, with_enter: bool = False) -> str:
    """Focus a window and type text into it. Set with_enter=True to press Enter after typing.
    Example: type_in_window(window_title_regex='Notepad', text='Hello World', with_enter=False)
    Example: type_in_window(window_title_regex='Run', text='notepad.exe', with_enter=True)"""
    return _tool_type_in_window(window_title_regex, text, with_enter)


@tool
def keyboard_type(text: str) -> str:
    """Type text at the current cursor position. Prefer type_in_window if you know the target window.
    Example: keyboard_type(text='Hello World')"""
    return _tool_keyboard_type(text)


@tool
def keyboard_press(keys: str) -> str:
    """Press a key or keyboard shortcut. Use '+' to combine modifier keys.
    Examples: keyboard_press(keys='enter'), keyboard_press(keys='ctrl+s'), keyboard_press(keys='alt+f4')"""
    return _tool_keyboard_press(keys)


@tool
def mouse_move(x: int, y: int, duration: float = 0.2) -> str:
    """Move the mouse pointer to screen coordinates (x, y).
    Example: mouse_move(x=500, y=300)"""
    return _tool_mouse_move(x, y, duration)


@tool
def mouse_click(button: str = "left", clicks: int = 1, x: int = -1, y: int = -1) -> str:
    """Click the mouse. button: 'left', 'right', or 'middle'. Use clicks=2 for double-click.
    Omit x/y to click at the current mouse position.
    Example: mouse_click(button='left', clicks=1, x=200, y=400)"""
    return _tool_mouse_click(button, clicks, x, y)


@tool
def take_screenshot(file_path: str = "") -> str:
    """Capture a screenshot. Saves to data/screenshots/ with a timestamp if file_path is not given.
    Example: take_screenshot(file_path='C:/Users/user/Desktop/screen.png')"""
    return _tool_take_screenshot(file_path)


@tool
def vision_act_on_screen(element_description: str, action: str = "click", type_text: str = "") -> str:
    """Visually find an element anywhere on the screen using a numbered grid and interact with it.
    Use this when you need to click buttons, icons, or fields in Desktop Apps (Spotify, WhatsApp, Settings, etc).
    DO NOT use this for standard web browsing, use browser_agent tools instead.
    
    Args:
        element_description: What to look for (e.g., 'Spotify Play button', 'WhatsApp Search bar', 'Cancel')
        action: 'click', 'double_click', 'right_click', or 'click_and_type'
        type_text: The text to type (ONLY used if action='click_and_type')
        
    Example: vision_act_on_screen(element_description='Windows Start Menu', action='click')
    Example: vision_act_on_screen(element_description='Search bar', action='click_and_type', type_text='Hello')
    """
    return _tool_vision_act_on_screen(element_description, action, type_text)


@tool
def play_media(query_or_path: str) -> str:
    """Play a media file by full path, filename keyword, or a generic request like 'any song'.
    Searches Music, Videos, Downloads, Desktop, and OneDrive folders automatically.
    If no specific file is requested, plays the first media file found on the system.
    Example: play_media(query_or_path='bohemian rhapsody')
    Example: play_media(query_or_path='any')  <- plays first available media file
    Example: play_media(query_or_path='C:/Music/song.mp3')"""
    return _tool_play_media(query_or_path)


@tool
def find_media_files(directory: str = "", extension_filter: str = "") -> str:
    """List media files in standard folders (Music, Videos, Downloads, Desktop, OneDrive).
    Optionally provide a specific directory or filter by extension (e.g. 'mp3,wav').
    Use this BEFORE play_media when the user asks to search for or list music/video files.
    Returns up to 50 files with full paths so the agent can then call play_media on one.
    Example: find_media_files()
    Example: find_media_files(directory='C:/Users/nchar/Downloads', extension_filter='mp3')"""
    return _tool_find_media_files(directory, extension_filter)



@tool
def web_search(query: str) -> str:
    """Open a Google web search in the default browser.
    Example: web_search(query='Python tutorials for beginners')"""
    return _tool_web_search(query)


@tool
def open_website(url: str) -> str:
    """Open a URL in the default browser. Prepends https:// if no scheme is provided.
    Example: open_website(url='https://www.github.com')
    Example: open_website(url='youtube.com')"""
    return _tool_open_website(url)


@tool
def create_file(file_path: str, content: str = "") -> str:
    """Create a new file or OVERWRITE an existing file with the given content.
    WARNING: destroys existing content. Use append_file to add to an existing file.
    Example: create_file(file_path='C:/Users/user/notes.txt', content='First line')"""
    return _tool_create_file(file_path, content)


@tool
def append_file(file_path: str, content: str) -> str:
    """Add text to the END of an existing file WITHOUT overwriting it.
    Use when user says 'append', 'add to', or 'write to' an existing file.
    The file must already exist; use create_file first if needed.
    Example: append_file(file_path='C:/notes.txt', content='\\nhello world')"""
    return _tool_append_file(file_path, content)


@tool
def read_file(file_path: str) -> str:
    """Read and return the text content of a file (max 4000 chars, then truncated).
    Example: read_file(file_path='C:/Users/user/Desktop/notes.txt')"""
    return _tool_read_file(file_path)


@tool
def delete_file(file_path: str) -> str:
    """Permanently delete a file. Cannot be undone.
    Example: delete_file(file_path='C:/Users/user/Desktop/old_notes.txt')"""
    return _tool_delete_file(file_path)


@tool
def copy_file(source_path: str, destination_path: str) -> str:
    """Copy a file from source to destination. Creates destination directories if needed.
    Example: copy_file(source_path='C:/docs/file.txt', destination_path='C:/backup/file.txt')"""
    return _tool_copy_file(source_path, destination_path)


@tool
def list_files(directory_path: str) -> str:
    """List files and subdirectories in a folder (up to 200 entries).
    Example: list_files(directory_path='C:/Users/user/Desktop')"""
    return _tool_list_files(directory_path)


@tool
def write_csv(file_path: str, headers_json: str = "", rows_json: str = "", append: bool = False) -> str:
    """Write CSV data.
    headers_json: JSON array of headers.
    rows_json: JSON array of rows (arrays or objects).
    Example headers_json='[\"name\",\"age\"]', rows_json='[[\"alice\",30],[\"bob\",28]]'."""
    return _tool_write_csv(file_path, headers_json, rows_json, append)


@tool
def read_csv(file_path: str, max_rows: int = 50) -> str:
    """Read CSV and return rows as JSON array."""
    return _tool_read_csv(file_path, max_rows)


@tool
def csv_to_excel(csv_path: str, excel_path: str, sheet_name: str = "Sheet1") -> str:
    """Convert CSV file to Excel (.xlsx)."""
    return _tool_csv_to_excel(csv_path, excel_path, sheet_name)


@tool
def excel_to_csv(excel_path: str, csv_path: str, sheet_name: str = "") -> str:
    """Convert Excel (.xlsx) to CSV. Uses first sheet if sheet_name not provided."""
    return _tool_excel_to_csv(excel_path, csv_path, sheet_name)


@tool
def list_excel_sheets(excel_path: str) -> str:
    """List sheet names in an Excel workbook."""
    return _tool_list_excel_sheets(excel_path)


@tool
def convert_image_format(input_path: str, output_format: str, output_path: str = "") -> str:
    """Convert image formats, e.g., jpg->png, png->jpeg, webp->png."""
    return _tool_convert_image_format(input_path, output_format, output_path)


@tool
def get_datetime() -> str:
    """Return the current local date and time. No arguments needed.
    Use when the user asks what time or date it is, or when a timestamp is needed.
    Example: get_datetime()"""
    return _tool_get_datetime()


@tool
def run_shell_command(command: str) -> str:
    """Run a Windows shell/cmd command and return its output. Timeout: 30 seconds.
    Use for system info, batch operations, or running scripts not covered by other tools.
    Example: run_shell_command(command='dir C:\\Users\\user\\Desktop')
    Example: run_shell_command(command='python C:\\scripts\\process.py')"""
    return _tool_run_shell_command(command)


@tool
def run_multi_step_actions(steps_json: str) -> str:
    """Execute a sequence of desktop actions from a JSON array. PREFER this for any task needing 2+ steps.
    Each item must have 'action' (string) and 'args' (object with named parameters).

    Available actions: open_application, open_file_with_app, close_application, focus_window,
    click_window_control, type_in_window, keyboard_type, keyboard_press, mouse_move, mouse_click,
    take_screenshot, vision_act_on_screen, find_media_files, play_media, web_search, open_website, create_file,
    append_file, read_file, delete_file, copy_file, list_files, write_csv, read_csv,
    csv_to_excel, excel_to_csv, list_excel_sheets, convert_image_format, get_datetime,
    run_shell_command, search_knowledge_base.

    Example — search for music then play one:
    [
      {"action": "find_media_files", "args": {}},
      {"action": "play_media", "args": {"query_or_path": "any"}}
    ]

    Example — open a file in notepad then append text to it:
    [
      {"action": "open_file_with_app", "args": {"file_path": "C:/Users/nchar/Desktop/prompt.txt", "app_name": "notepad"}},
      {"action": "append_file", "args": {"file_path": "C:/Users/nchar/Desktop/prompt.txt", "content": "\\nhello world"}}
    ]

    Example — open notepad, wait is not needed, just type after opening:
    [
      {"action": "open_application", "args": {"app_name": "notepad"}},
      {"action": "type_in_window", "args": {"window_title_regex": "Notepad", "text": "Hello World"}}
    ]

    Example — get current time:
    [{"action": "get_datetime", "args": {}}]
    """
    return _tool_run_multi_step_actions(steps_json)


# ── Ask user for clarification ────────────────────────────────────────────────

# Pluggable input handler — defaults to terminal input().
# GUI frontends (PyQt5, etc.) can replace this with a thread-safe dialog.
# Signature: (context: str, question: str) -> str
_ASK_USER_HANDLER = None


def set_ask_user_handler(handler) -> None:
    """Register a custom input handler for the ask_user tool.
    The handler receives (context: str, question: str) and must return a string answer.
    If not set, falls back to blocking terminal input().
    """
    global _ASK_USER_HANDLER
    _ASK_USER_HANDLER = handler


def _tool_ask_user(question: str, context: str = "") -> str:
    """Pause and ask the human user a clarifying question, then return their answer."""
    handler = _ASK_USER_HANDLER

    if handler is not None:
        try:
            return str(handler(context, question))
        except Exception as exc:
            return f"[ask_user handler error: {exc}]"

    # ── Default: terminal prompt ──────────────────────────────────────────────
    sep = "─" * 60
    print(f"\n{sep}")
    if context.strip():
        print(f"[Sentinel] Context:\n{context.strip()}")
        print(sep)
    print(f"[Sentinel] {question}")
    print(sep)
    try:
        answer = input("Your answer: ").strip()
    except (EOFError, KeyboardInterrupt):
        answer = ""
    print(sep + "\n")
    return answer if answer else "(no answer provided)"


@tool
def ask_user(question: str, context: str = "") -> str:
    """Ask the human user a clarifying question and return their answer before continuing.

    Use this tool when:
    - The user's request is ambiguous and you need more information to proceed
    - There are multiple valid ways to complete a task and you need the user to choose
    - A destructive action (delete, overwrite) needs explicit confirmation
    - You have partial results and need the user to pick a direction

    Args:
        question: The specific question to ask the user.
        context:  Optional summary of what has happened so far, so the user
                  understands why you are asking. Include relevant facts,
                  tool results, or the options available.

    Returns the user's answer as a string.

    Examples:
        ask_user(
            question="Which file should I open — 'report_v1.docx' or 'report_final.docx'?",
            context="I found two files matching 'report' on the Desktop."
        )
        ask_user(
            question="Should I delete all 12 log files, or only the ones older than 7 days?",
            context="Found 12 .log files in C:/Logs. 5 are older than 7 days."
        )
        ask_user(
            question="The music folder is empty. Should I search Downloads instead?",
            context="No media files found in Music or Videos folders."
        )
    """
    return _tool_ask_user(question, context)


@tool
def open_folder(folder_path: str) -> str:
    """Open a folder in Windows File Explorer.
    Accepts: full paths, shortcuts (desktop, downloads, onedrive),
    or relative paths like OneDrive/Desktop, Desktop/myproject.
    Example: open_folder('OneDrive/Desktop')
    Example: open_folder('desktop')
    Example: open_folder('C:/Projects/myapp')"""
    import os as _os

    home = Path.home()
    real_desktop   = _resolve_special_folder("desktop")
    real_documents = _resolve_special_folder("documents")

    # Find real OneDrive folder (handles OneDrive - Personal, regional names)
    onedrive_env = _os.getenv("OneDrive", "")
    onedrive = Path(onedrive_env) if onedrive_env and Path(onedrive_env).exists() else None
    if onedrive is None:
        for d in home.iterdir():
            if d.is_dir() and d.name.lower().startswith("onedrive"):
                onedrive = d
                break
    if onedrive is None:
        onedrive = home / "OneDrive"

    shortcuts = {
        "desktop":             real_desktop,
        "downloads":           home / "Downloads",
        "documents":           real_documents,
        "pictures":            _resolve_special_folder("pictures"),
        "music":               _resolve_special_folder("music"),
        "videos":              _resolve_special_folder("videos"),
        "onedrive":            onedrive,
        "onedrive/desktop":    onedrive / "Desktop",
        "onedrive\\desktop": onedrive / "Desktop",
        "home":                home,
    }

    key = folder_path.strip().lower().rstrip("/\\").replace("\\", "/")
    resolved = shortcuts.get(key)

    if resolved is None:
        p_norm = folder_path.replace("\\", "/")
        p_lower = p_norm.lower()

        if p_lower.startswith("onedrive/"):
            suffix = p_norm[len("OneDrive/"):]
            resolved = onedrive / suffix
        elif not Path(folder_path).is_absolute():
            # Try real Desktop, OneDrive/Desktop, home
            for base in [real_desktop, onedrive / "Desktop", home]:
                candidate = base / folder_path
                if candidate.exists():
                    resolved = candidate
                    break
            if resolved is None:
                resolved = home / folder_path
        else:
            resolved = _expand_path(folder_path)

    resolved = Path(resolved)

    if not resolved.exists():
        return (
            f"Folder not found: {resolved}\n"
            f"OneDrive root detected: {onedrive}\n"
            f"Tip: use full path e.g. {onedrive / 'Desktop'}"
        )

    # os.startfile is correct Windows API — handles spaces in path natively
    try:
        os.startfile(str(resolved))
        time.sleep(1.2)
        return f"Opened folder in Explorer: {resolved}"
    except Exception:
        try:
            subprocess.Popen(f'explorer.exe "{resolved}"', shell=True)
            time.sleep(1.2)
            return f"Opened folder in Explorer: {resolved}"
        except Exception as exc:
            return f"Failed to open folder '{resolved}': {exc}"


@tool
def open_cmd_and_run(command: str = "", working_dir: str = "", keep_open: bool = True) -> str:
    """Open a CMD (Command Prompt) window and optionally run a command in it.
    Use this when the user says 'open cmd', 'run cmd', 'open terminal', or wants
    a visible CMD window rather than a silent background command.
    - command: the command to run inside CMD (leave empty to just open CMD)
    - working_dir: directory to start in (default: user's home)
    - keep_open: if True, CMD stays open after the command runs (default True)
    Example: open_cmd_and_run(command='python script.py', working_dir='C:/Projects')
    Example: open_cmd_and_run(command='', keep_open=True)  <- just opens CMD
    Example: open_cmd_and_run(command='dir', keep_open=True)"""
    cwd = str(Path(working_dir).expanduser()) if working_dir.strip() else str(Path.home())
    try:
        if not command.strip():
            # Just open CMD
            subprocess.Popen(["cmd.exe"], cwd=cwd,
                             creationflags=subprocess.CREATE_NEW_CONSOLE)
            time.sleep(0.8)
            return "Opened CMD window."
        if keep_open:
            # /K keeps the window open after command
            subprocess.Popen(
                ["cmd.exe", "/K", command],
                cwd=cwd, creationflags=subprocess.CREATE_NEW_CONSOLE
            )
        else:
            subprocess.Popen(
                ["cmd.exe", "/C", command],
                cwd=cwd, creationflags=subprocess.CREATE_NEW_CONSOLE
            )
        time.sleep(1.0)
        return f"Opened CMD and ran: {command}"
    except Exception as exc:
        return f"Failed to open CMD: {exc}"


@tool
def set_clipboard(text: str) -> str:
    """Copy text to the Windows clipboard. Useful before pasting into an app
    that does not support direct typing (e.g. WhatsApp Desktop search box, some web inputs).
    Follow with keyboard_press(keys='ctrl+v') to paste.
    Example: set_clipboard(text='Hello, how are you?')"""
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             f"Set-Clipboard -Value '{text.replace(chr(39), chr(39)+chr(39))}'"],
            capture_output=True, timeout=5
        )
        return f"Copied to clipboard: {text[:80]}{'...' if len(text)>80 else ''}"
    except Exception as exc:
        return f"Clipboard set failed: {exc}"


@tool
def click_element_by_text(window_title_regex: str, text: str, control_type: str = "") -> str:
    """Find and click a UI element inside a window by its visible text label.
    More reliable than click_window_control for buttons, links, and menu items
    where you know the text but not the exact control hierarchy.
    - window_title_regex: regex to match the window title
    - text: exact visible text of the element to click
    - control_type: optional filter — 'Button', 'MenuItem', 'Hyperlink', 'Edit', etc.
    Example: click_element_by_text(window_title_regex='WhatsApp', text='New chat')
    Example: click_element_by_text(window_title_regex='Antigravity.*', text='Send', control_type='Button')
    Example: click_element_by_text(window_title_regex='.*Chrome.*', text='Address bar')"""
    try:
        _, Desktop, _ = _get_pywinauto()
        # Fast Win32 lookup — never hangs on VSCode/Chrome/Electron apps
        _handles = _get_windows_by_title(window_title_regex)
        if not _handles:
            return f"No window found matching: '{window_title_regex}'"
        win = _FastWindow(max(_handles))
        win.set_focus()
        time.sleep(0.3)
        kwargs = {"title": text}
        if control_type.strip():
            kwargs["control_type"] = control_type
        ctrl = win.child_window(**kwargs)
        ctrl.wrapper_object().click_input()
        return f"Clicked '{text}' in '{win.window_text()}'."
    except Exception as exc:
        return f"Could not click '{text}' in '{window_title_regex}': {exc}"


@tool
def search_files(
    name_pattern: str = "",
    search_dir: str = "",
    extension: str = "",
    content_keyword: str = "",
    min_size_kb: float = -1,
    max_size_kb: float = -1,
    max_results: int = 50,
) -> str:
    """Recursively search for files or folders. All filters are optional and combinable.

    Args:
        name_pattern:    Substring to match against file/folder name (case-insensitive).
                         Supports wildcards: 'report*', '*.log'.
                         Leave empty to match everything.
        search_dir:      Root directory to search in (default: user home folder C:/Users/nchar).
        extension:       File extension filter — e.g. 'txt', 'pdf', 'py', 'docx'.
                         Do NOT include the dot.
        content_keyword: Search INSIDE text files for this keyword. Empty = skip content scan.
        min_size_kb:     Minimum file size in KB (-1 = no limit).
        max_size_kb:     Maximum file size in KB (-1 = no limit).
        max_results:     Maximum results to return (default 50).

    Returns a numbered list of matching absolute paths.

    Examples:
        search_files(name_pattern='report', search_dir='C:/Users/nchar/Documents')
        search_files(extension='pdf', search_dir='C:/Users/nchar/Downloads')
        search_files(name_pattern='budget', extension='xlsx')
        search_files(content_keyword='invoice', search_dir='C:/Projects', extension='txt')
        search_files(extension='log', min_size_kb=100, max_size_kb=5000)
        search_files(name_pattern='screenshot', extension='png', search_dir='C:/Users/nchar/Desktop')
    """
    return _tool_search_files(
        name_pattern=name_pattern,
        search_dir=search_dir,
        extension=extension,
        content_keyword=content_keyword,
        min_size_kb=min_size_kb,
        max_size_kb=max_size_kb,
        max_results=max_results,
    )


@tool
def get_file_info(path: str) -> str:
    """Get detailed metadata and properties of a file or folder.

    Returns:
      - Full path, type (file/directory), name, extension, parent folder
      - Created, modified, and accessed timestamps
      - File size in human-readable form (B / KB / MB / GB)
      - MIME type and encoding (UTF-8 text or binary) for files
      - Number of immediate children and total size for folders
      - Read / write permissions

    Args:
        path: Full path to any file or folder.

    Examples:
        get_file_info(path='C:/Users/nchar/Desktop/report.pdf')
        get_file_info(path='C:/Projects/myapp')
        get_file_info(path='C:/Users/nchar/Documents')
    """
    return _tool_get_file_info(path)


@tool
def list_folder_tree(
    directory_path: str,
    max_depth: int = 3,
    max_items: int = 200,
) -> str:
    """Show an indented tree of a folder's contents (like the `tree` command).

    Useful when the user wants to see the full structure of a project or directory.
    Directories come first, files show their size in parentheses.

    Args:
        directory_path: Full path to the folder.
        max_depth:      How many levels deep to expand (default 3).
        max_items:      Max total items to display (default 200).

    Examples:
        list_folder_tree(directory_path='C:/Users/nchar/Desktop')
        list_folder_tree(directory_path='C:/Projects/myapp', max_depth=5)
        list_folder_tree(directory_path='C:/Users/nchar/Downloads', max_depth=2, max_items=100)
    """
    return _tool_list_folder_tree(directory_path, max_depth, max_items)


@tool
def scroll_window(window_title_regex: str, direction: str = "down", amount: int = 3) -> str:
    """Scroll inside a window up or down by a number of clicks.
    Useful for finding elements that are not visible without scrolling.
    - direction: 'up' or 'down'
    - amount: number of scroll clicks (default 3)
    Example: scroll_window(window_title_regex='WhatsApp', direction='down', amount=5)"""
    try:
        pyautogui = _get_pyautogui()
        _, Desktop, _ = _get_pywinauto()
        # Fast Win32 lookup — never hangs on VSCode/Chrome/Electron apps
        _handles = _get_windows_by_title(window_title_regex)
        if not _handles:
            return f"No window found matching: '{window_title_regex}'"
        win = _FastWindow(max(_handles))
        win.set_focus()
        time.sleep(0.2)
        rect = win.rectangle()
        cx = (rect.left + rect.right) // 2
        cy = (rect.top + rect.bottom) // 2
        pyautogui.moveTo(cx, cy)
        clicks = amount if direction.lower() == "up" else -amount
        pyautogui.scroll(clicks)
        return f"Scrolled {direction} {amount} in '{win.window_text()}'."
    except Exception as exc:
        return f"Scroll failed in '{window_title_regex}': {exc}"


@tool
def get_desktop_path() -> str:
    """Return the exact, true path of the current user's Desktop.
    Use this if you need to know exactly where the Desktop is, as it can be redirected
    (e.g., to OneDrive like C:/Users/nchar/OneDrive/Desktop)."""
    return str(_resolve_special_folder("desktop"))



def _tool_open_in_app(app_name: str, path: str, wait: float = 2.0) -> str:
    """
    Open a file OR folder in a specific application.
    Handles VSCode (code . ), explorer, notepad, etc.
    Uses app registry for exe path, subprocess for launch.
    """
    from app.src.app_registry import registry as _reg
    import os

    # Resolve path — handles OneDrive/Desktop, Desktop/x, relative paths
    resolved = path.strip()
    home = os.path.expanduser("~")

    # Find actual OneDrive folder (handles "OneDrive - Personal", regional variants)
    onedrive = os.getenv("OneDrive", os.path.join(home, "OneDrive"))
    if not os.path.exists(onedrive):
        for d in os.listdir(home):
            if d.lower().startswith("onedrive") and os.path.isdir(os.path.join(home, d)):
                onedrive = os.path.join(home, d)
                break

    # Normalize separators
    resolved_norm = resolved.lower().replace("\\", "/")

    if resolved_norm.startswith("onedrive/") or resolved_norm == "onedrive":
        suffix = resolved[len("OneDrive/"):] if "/" in resolved else ""
        resolved = os.path.join(onedrive, suffix) if suffix else onedrive
    elif resolved_norm.startswith("desktop/") or resolved_norm.startswith("desktop\\"):
        real_desktop = str(_resolve_special_folder("desktop"))
        suffix = resolved.split("/", 1)[-1].split("\\", 1)[-1]
        resolved = os.path.join(real_desktop, suffix)
    elif resolved_norm == "desktop":
        resolved = str(_resolve_special_folder("desktop"))
    elif not os.path.isabs(resolved):
        # Try real Desktop, OneDrive/Desktop, home in order
        real_desktop = str(_resolve_special_folder("desktop"))
        for base in [real_desktop, os.path.join(onedrive, "Desktop"), home]:
            candidate = os.path.join(base, resolved)
            if os.path.exists(candidate):
                resolved = candidate
                break

    if not os.path.exists(resolved):
        return f"Path not found: '{resolved}' (tried: {path})"

    exe = _reg.get(app_name) or _resolve_app_command(app_name)
    if not exe:
        return f"App not found: {app_name}"

    try:
        # Quoted shell command handles spaces in both exe path and folder path
        # e.g. "C:/Program Files/Code.exe" "C:/Users/nchar/OneDrive/Desktop"
        cmd = f'"{exe}" "{resolved}"'
        subprocess.Popen(cmd, shell=True,
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(wait)
        return f"Opened '{resolved}' in {app_name} ({exe})"
    except Exception as exc:
        try:
            os.startfile(str(resolved))
            return f"Opened '{resolved}' with system default"
        except Exception:
            return f"Failed: {exc}"


@tool
def open_in_app(app_name: str, path: str) -> str:
    """Open a file or folder in a specific application.

    Better than open_file_with_app for opening folders in editors like VSCode.
    Resolves Desktop/ shortcuts automatically.

    Args:
        app_name: Application to use ('vscode', 'notepad', 'explorer', 'vlc', etc.)
        path: File or folder path. Supports:
              - Absolute: C:/Users/nchar/Projects/myapp
              - Desktop shortcut: Desktop/my_project
              - Relative to Desktop: new_sentinel

    Examples:
        open_in_app("vscode", "Desktop/new_sentinel")
        open_in_app("vscode", "C:/Users/nchar/Desktop/new_sentinel")
        open_in_app("explorer", "Desktop/Downloads")
        open_in_app("notepad", "C:/Users/nchar/Desktop/notes.txt")
    """
    return _tool_open_in_app(app_name, path)



@tool
def manage_custom_tool(action: str, tool_name: str, python_code: str = "") -> str:
    """Add, edit, or delete a custom tool dynamically.
    
    Args:
        action: 'add', 'edit', or 'delete'
        tool_name: The name of the tool (e.g., 'scrape_weather_api')
        python_code: The complete, executable Python code for the tool, including imports and the @tool decorator.
        
    Example: manage_custom_tool('add', 'my_tool', 'from langchain_core.tools import tool\\n\\n@tool\\ndef my_tool():\\n    return "hi"')
    """
    import os
    
    # Define a safe directory for dynamic tools
    custom_tools_dir = Path.cwd() / "app" / "src" / "custom_tools"
    custom_tools_dir.mkdir(parents=True, exist_ok=True)
    
    # Ensure __init__.py exists so it acts as a module
    init_file = custom_tools_dir / "__init__.py"
    if not init_file.exists():
        init_file.write_text("")

    file_path = custom_tools_dir / f"{tool_name}.py"
    
    if action.lower() in ['add', 'edit']:
        if not python_code.strip():
            return "Error: python_code cannot be empty for add/edit."
        
        try:
            # Write the code
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(python_code)
            
            # Syntax check before confirming
            import ast
            ast.parse(python_code)
            
            return f"SUCCESS: Custom tool '{tool_name}' has been {action}ed at {file_path}. Sentinel can now use this tool on the next orchestrator run."
            
        except SyntaxError as e:
            # Rollback if syntax is garbage
            if file_path.exists():
                file_path.unlink()
            return f"SYNTAX ERROR in provided code: {e}. Tool was NOT saved."
            
    elif action.lower() == 'delete':
        if file_path.exists():
            file_path.unlink()
            return f"SUCCESS: Custom tool '{tool_name}' has been deleted."
        return f"Error: Tool '{tool_name}' not found."
        
    return "Error: Action must be 'add', 'edit', or 'delete'."
    
def get_tools():
    return [
        search_knowledge_base,
        ask_user,
        open_application,
        open_in_app,
        open_folder,
        open_file_with_app,
        close_application,
        kill_app_instances,
        sleep,
        focus_window,
        click_window_control,
        click_element_by_text,
        type_in_window,
        set_clipboard,
        keyboard_type,
        keyboard_press,
        mouse_move,
        mouse_click,
        scroll_window,
        take_screenshot,
        vision_act_on_screen,
        play_media,
        find_media_files,
        web_search,
        open_website,
        create_file,
        append_file,
        read_file,
        delete_file,
        copy_file,
        list_files,
        write_csv,
        read_csv,
        csv_to_excel,
        excel_to_csv,
        list_excel_sheets,
        convert_image_format,
        get_datetime,
        run_shell_command,
        open_cmd_and_run,
        search_files,
        get_file_info,
        list_folder_tree,
        get_desktop_path,
        run_multi_step_actions,
        manage_custom_tool,
    ]